import openpyxl
wb = openpyxl.load_workbook(r'C:\Users\jakep\OneDrive\Documents\Port Authority\2026\port-finances-26\finance_data.xlsx')

print('=== Player Balances B7 (Total Charged) — full formula ===')
ws = wb['Player Balances']
print(ws['B7'].value)

print()
print('=== Attendance row 4 col B (cost per person) — full formula ===')
ws = wb['Attendance']
print(ws.cell(4, 2).value)

print()
print('=== Attendance row 5 col B (total attendees) ===')
print(ws.cell(5, 2).value)

print()
print('=== Events col G rows 4-12 (Actual Cost) ===')
ws = wb['Events']
for r in range(4, 13):
    print(f'  G{r}: {ws.cell(r, 7).value}')
