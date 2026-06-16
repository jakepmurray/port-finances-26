import openpyxl
wb = openpyxl.load_workbook(r'C:\Users\jakep\OneDrive\Documents\Port Authority\2026\port-finances-26\finance_data.xlsx')

print('=== Player Balances B7 ===')
ws = wb['Player Balances']
v = str(ws['B7'].value)
print(v[:160])
if '="1"' in v:
    print('FAIL - still has string comparison ="1"')
elif '=1)' in v or '=1)*' in v:
    print('PASS - numeric =1 comparison present')
else:
    print('UNKNOWN - check manually')

print()
print('=== Events col G rows 5-7 ===')
ws = wb['Events']
for r in [5, 6, 7]:
    val = str(ws.cell(r,7).value or '')
    print(f'  G{r}: {val[:90]}')

print()
print('=== Attendance B4 (still references Events G5) ===')
ws = wb['Attendance']
print(str(ws.cell(4,2).value)[:120])
