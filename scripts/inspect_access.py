import openpyxl
wb = openpyxl.load_workbook(r'C:\Users\jakep\OneDrive\Documents\Port Authority\2026\port-finances-26\finance_data.xlsx')

print('=== Sheets ===')
print(wb.sheetnames)

print('\n=== Access Codes sheet (first 10 rows) ===')
ws = wb['Access Codes']
for r in range(1, 11):
    row = [ws.cell(r, c).value for c in range(1, 6)]
    if any(v is not None for v in row):
        print(f'  row{r}: {row}')

print('\n=== Roster sheet (first 8 rows, first 8 cols) ===')
ws = wb['Roster']
for r in range(1, 9):
    row = [ws.cell(r, c).value for c in range(1, 9)]
    if any(v is not None for v in row):
        print(f'  row{r}: {row}')
