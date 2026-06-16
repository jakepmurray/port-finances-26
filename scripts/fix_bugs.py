import shutil
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

SRC = r'C:\Users\jakep\OneDrive\Documents\Port Authority\2026\port-finances-26\finance_data.xlsx'
TMP = r'C:\Users\jakep\AppData\Local\Temp\finance_data_bugfix.xlsx'
shutil.copy(SRC, TMP)

wb = load_workbook(TMP)

GREEN = '00008000'
thin = Side(style='thin')
BORDER = Border(top=thin, bottom=thin, left=thin, right=thin)

# ── FIX 1: Player Balances B7:B206 — change ="1" to =1 ───────────────
# The SUMPRODUCT comparison ="1" is a strict string match; users entering
# numeric 1 in Attendance cells won't match. Changing to =1 (numeric)
# lets Excel/Sheets coerce both "1" and 1 correctly.
ws_pb = wb['Player Balances']
fixed_pb = 0
for r in range(7, 207):
    cell = ws_pb.cell(r, 2)
    if cell.value and '="1"' in str(cell.value):
        cell.value = cell.value.replace('="1"', '=1')
        fixed_pb += 1
print(f'Player Balances: updated {fixed_pb} Total Charged formulas ("=1" string -> numeric =1)')

# ── FIX 2: Events col G — formula-driven from Transactions ───────────
# Replace hardcoded Actual Cost values with SUMIFS that totals all
# Expense-type transactions linked to this event. This makes the full
# chain live: Transactions → Events Actual Cost → Attendance row 4
# (cost per person) → Player Balances Total Charged.
ws_ev = wb['Events']
for r in range(5, 505):
    cell = ws_ev.cell(r, 7)
    cell.value = (
        f'=IFERROR(SUMIFS(Transactions!$F$7:$F$1000,'
        f'Transactions!$G$7:$G$1000,A{r},'
        f'Transactions!$D$7:$D$1000,"Expense"),0)'
    )
    # Re-style font to green (cross-sheet formula); preserve existing fill/border
    existing = cell.font
    cell.font = Font(
        name=existing.name or 'Arial',
        size=existing.size or 10,
        bold=existing.bold,
        color=GREEN,
    )
    cell.alignment = Alignment(horizontal='right', vertical='center')
    cell.border = BORDER
print('Events: replaced col G (Actual Cost) rows 5-504 with SUMIFS formula')

wb.save(TMP)
shutil.copy(TMP, SRC)
print(f'\nSaved: {SRC}')
