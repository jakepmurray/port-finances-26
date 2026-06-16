import openpyxl
import sys

path = sys.argv[1]
wb = openpyxl.load_workbook(path)
print('Sheets:', wb.sheetnames)
for name in wb.sheetnames:
    ws = wb[name]
    print(f'\n=== {name} (max_row={ws.max_row}, max_col={ws.max_column}) ===')
    for row in ws.iter_rows(min_row=1, max_row=min(20, ws.max_row), values_only=True):
        safe = tuple(str(c).encode('ascii','replace').decode() if c is not None else None for c in row)
        print(safe)
