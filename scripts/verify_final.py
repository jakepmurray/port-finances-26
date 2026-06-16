import openpyxl
wb = openpyxl.load_workbook(r'C:\Users\jakep\AppData\Local\Temp\finance_data_final.xlsx')

print('=== Attendance row 4 (cost per person formulas, first 4 event cols) ===')
ws = wb['Attendance']
for c in range(1, 6):
    v = ws.cell(4, c).value
    print(f'  col{c}: {str(v)[:90] if v else "-"}')

print()
print('=== Transactions col J (Cost per Attendee) rows 6-14 ===')
ws = wb['Transactions']
for r in range(6, 15):
    g = str(ws.cell(r, 7).value or '-')[:25]
    j = str(ws.cell(r, 10).value or '-')[:85]
    print(f'  row{r}  G={g:25s}  J={j}')

print()
print('=== Player Balances row 7 (Total Charged formula) ===')
ws = wb['Player Balances']
print(f'  B7: {str(ws["B7"].value)[:120]}')
