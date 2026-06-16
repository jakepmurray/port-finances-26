import shutil
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

SRC = r'C:\Users\jakep\AppData\Local\Temp\finance_data_copy.xlsx'
OUT = r'C:\Users\jakep\AppData\Local\Temp\finance_data_modified.xlsx'
shutil.copy(SRC, OUT)

wb = load_workbook(OUT)

# =========================================================
# 1. ROSTER SHEET
# =========================================================
ws_r = wb['Roster']

# Delete summary stats rows 3-4
ws_r.delete_rows(3, 2)
# Now: row1=title, row2=blank, row3=blank(was5), row4=headers(was6), row5+=data

# Delete col B (Access Code), then col B again (Jersey #), then col B again (Date Joined)
# Each delete shifts columns left
ws_r.delete_cols(2)  # removes Access Code
ws_r.delete_cols(2)  # removes Jersey # (was C)
ws_r.delete_cols(2)  # removes Date Joined (was D)
# Remaining cols: A=Full Name, B=Dues Amount, C=Dues Paid?, D=Role, E=Status, F=Notes

# Update headers
ws_r['A4'] = 'Full Name'
ws_r['B4'] = 'Dues Amount ($)'
ws_r['C4'] = 'Dues Paid?'
ws_r['D4'] = 'Role'
ws_r['E4'] = 'Status'
ws_r['F4'] = 'Notes'
# Clear title row 1 (had "ROSTER & ACCESS CODES") - update it
ws_r['A1'] = 'ROSTER - Ultimate Frisbee Team'

# =========================================================
# 2. EVENTS SHEET
# =========================================================
ws_e = wb['Events']

# Delete summary stats rows 3-4
ws_e.delete_rows(3, 2)
# Now: row4=headers(was6), row5+=data

# Current headers row4: A=Event Name, B=Date, C=Category, D=Location, E=# Expected,
#                        F=Budgeted, G=Actual, H=Who Paid Upfront, I=Split Method, J=Notes

# Rename Date -> Start Date and insert End Date as col C
ws_e['B4'] = 'Start Date'
ws_e.insert_cols(3)  # inserts blank col C, pushes Category+ right
ws_e['C4'] = 'End Date'
# Now: A=Event Name, B=Start Date, C=End Date, D=Category, E=Location,
#      F=# Expected, G=Budgeted, H=Actual, I=Who Paid, J=Split, K=Notes

# Delete col F (# Expected)
ws_e.delete_cols(6)
# Now: A=Event Name, B=Start Date, C=End Date, D=Category, E=Location,
#      F=Budgeted Cost ($), G=Actual Cost ($), H=Who Paid Upfront, I=Split Method, J=Notes

ws_e['A1'] = 'EVENTS MASTER - All Team Events & Costs'

# =========================================================
# 3. ATTENDANCE SHEET - full rebuild
# =========================================================
ws_att = wb['Attendance']
ws_att.delete_rows(1, ws_att.max_row)

# Row 1: title
ws_att['A1'] = 'ATTENDANCE TRACKER'

# Row 3: column headers - Player Name + event names from Events (50 events, cols B-AY)
ws_att['A3'] = 'Player Name'
for i in range(50):
    col = get_column_letter(i + 2)  # B through AY
    event_row = 5 + i               # Events data starts at row 5
    ws_att[f'{col}3'] = f'=IF(Events!$A${event_row}<>"",Events!$A${event_row},"")'

# Row 4: Cost per person (manual input per event, labeled)
ws_att['A4'] = 'Cost per person ($)'

# Row 5: Total Attendees (COUNTIF per event column)
ws_att['A5'] = 'Total Attendees'
for i in range(50):
    col = get_column_letter(i + 2)
    ws_att[f'{col}5'] = f'=COUNTIF({col}7:{col}206,"1")'

# Row 6: Total Cost per event
ws_att['A6'] = 'Total Cost ($)'
for i in range(50):
    col = get_column_letter(i + 2)
    ws_att[f'{col}6'] = f'={col}4*{col}5'

# Rows 7-206: Player names from Roster (Roster data rows 5-204)
for j in range(200):
    att_row = 7 + j
    roster_row = 5 + j
    ws_att[f'A{att_row}'] = f'=IF(Roster!$A${roster_row}<>"",Roster!$A${roster_row},"")'

# =========================================================
# 4. TRANSACTIONS - Linked Events dropdown
# =========================================================
ws_t = wb['Transactions']
# Events data starts at row 5 after our changes to Events sheet
dv = DataValidation(
    type="list",
    formula1="Events!$A$5:$A$505",
    allow_blank=True,
    showDropDown=False
)
ws_t.add_data_validation(dv)
dv.sqref = "G7:G1000"

# =========================================================
# 5. PLAYER BALANCES - restructure
# =========================================================
ws_pb = wb['Player Balances']

# Update title
ws_pb['A1'] = 'PLAYER BALANCES - Individual Ledger'

# Clear and reset summary row headers (row 3)
ws_pb['A3'] = 'Total Owed TO Team'
ws_pb['B3'] = 'Total Owed BY Team'
ws_pb['C3'] = 'Net'
ws_pb['D3'] = 'Settled'
ws_pb['E3'] = 'Unsettled'
# Clear extra header cells
for col in range(6, 9):
    ws_pb.cell(row=3, column=col).value = None

# Summary formulas row 4 - will reference new col E (Net Balance) and col F (Status)
# We'll update these after restructuring columns

# Set new column headers row 6
ws_pb['A6'] = 'Player Name'
ws_pb['B6'] = 'Total Charged ($)'
ws_pb['C6'] = 'Total Paid ($)'
ws_pb['D6'] = 'Paid Upfront For Team ($)'
ws_pb['E6'] = 'Net Balance ($)'
ws_pb['F6'] = 'Status'
# Clear old columns G, H
ws_pb['G6'] = None
ws_pb['H6'] = None

# Clear all existing data in rows 7-206 for cols A-H
for row in range(7, 207):
    for col in range(1, 9):
        ws_pb.cell(row=row, column=col).value = None

# Populate formulas for rows 7-206
for j in range(200):
    r = 7 + j
    roster_r = 5 + j

    # A: Player Name from Roster
    ws_pb[f'A{r}'] = f'=IF(Roster!$A${roster_r}<>"",Roster!$A${roster_r},"")'

    # B: Total Charged = dues from Roster + event costs from Attendance
    dues = f'IFERROR(INDEX(Roster!$B$5:$B$204,MATCH(A{r},Roster!$A$5:$A$204,0)),0)'
    events = (
        f'IFERROR(SUMPRODUCT('
        f'(INDEX(Attendance!$B$7:$AY$206,MATCH(A{r},Attendance!$A$7:$A$206,0),0)="1")'
        f'*Attendance!$B$4:$AY$4),0)'
    )
    ws_pb[f'B{r}'] = f'=IF(A{r}<>"",{dues}+{events},"")'

    # C: Total Paid = income transactions by this player
    ws_pb[f'C{r}'] = (
        f'=IF(A{r}<>"",'
        f'SUMIFS(Transactions!$F$7:$F$1000,'
        f'Transactions!$D$7:$D$1000,"Income",'
        f'Transactions!$E$7:$E$1000,A{r}),"")'
    )

    # D: Paid Upfront = expense transactions paid by this player
    ws_pb[f'D{r}'] = (
        f'=IF(A{r}<>"",'
        f'SUMIFS(Transactions!$F$7:$F$1000,'
        f'Transactions!$D$7:$D$1000,"Expense",'
        f'Transactions!$E$7:$E$1000,A{r}),"")'
    )

    # E: Net Balance = Total Charged - Total Paid - Paid Upfront
    ws_pb[f'E{r}'] = f'=IF(A{r}<>"",B{r}-C{r}-D{r},"")'

    # F: Status
    ws_pb[f'F{r}'] = f'=IF(A{r}<>"",IF(E{r}=0,"Settled","Unsettled"),"")'

# Update summary row 4 to reference new column positions
ws_pb['A4'] = '=SUMIF(E7:E206,">0",E7:E206)'
ws_pb['B4'] = '=SUMIF(E7:E206,"<0",E7:E206)'
ws_pb['C4'] = '=SUM(E7:E206)'
ws_pb['D4'] = '=COUNTIF(F7:F206,"Settled")'
ws_pb['E4'] = '=COUNTIF(F7:F206,"Unsettled")'
# Clear old summary cells F-H row 4
for col in range(6, 9):
    ws_pb.cell(row=4, column=col).value = None

# =========================================================
# 6. DASHBOARD - update broken references
# =========================================================
ws_dash = wb['Dashboard']
# Roster summary rows were deleted; recalculate from raw data
# Roster after changes: data rows 5-204, cols A=Name B=Dues C=DuesPaid? D=Role E=Status
ws_dash['B4'] = '=COUNTA(Roster!A5:A204)'
ws_dash['B5'] = '=COUNTIF(Roster!E5:E204,"Active")'
ws_dash['B6'] = '=COUNTIF(Roster!D5:D204,"Coach")+COUNTIF(Roster!D5:D204,"Treasurer")'
ws_dash['B7'] = '=COUNTIF(Roster!D5:D204,"Captain")'
ws_dash['B10'] = '=SUMIF(Roster!E5:E204,"Active",Roster!B5:B204)'
ws_dash['B11'] = '=SUMIFS(Transactions!F7:F1000,Transactions!D7:D1000,"Income",Transactions!C7:C1000,"Dues")'
ws_dash['B12'] = '=B10-B11'
# Events after changes: data rows 5-505, col D=Category
ws_dash['B19'] = '=COUNTA(Events!A5:A505)'
ws_dash['B20'] = '=COUNTIF(Events!D5:D505,"Practice")'
# Check if there are more rows in Dashboard for other event types
for row in range(21, 30):
    cell = ws_dash.cell(row=row, column=1).value
    val = ws_dash.cell(row=row, column=2).value
    if val and 'Events!' in str(val):
        # Update Events category references: old col C -> new col D
        new_val = str(val).replace('Events!C', 'Events!D').replace('Events!A4', 'COUNTA(Events!A5:A505)')
        # Also fix row offset: old row 4 was summary, now direct COUNTIF
        ws_dash.cell(row=row, column=2).value = new_val

# =========================================================
# 7. PENDING APPROVAL tab (new sheet)
# =========================================================
# Insert before Dashboard
dash_idx = wb.sheetnames.index('Dashboard')
ws_pend = wb.create_sheet('Pending Approval', dash_idx)
ws_pend['A1'] = 'PENDING TRANSACTIONS - Player Submissions'
ws_pend['A2'] = 'Players submit items here. Treasurer reviews and moves approved items to the Transactions sheet.'
ws_pend['A4'] = 'Date'
ws_pend['B4'] = 'Player Name'
ws_pend['C4'] = 'Description'
ws_pend['D4'] = 'Category'
ws_pend['E4'] = 'Amount ($)'
ws_pend['F4'] = 'Receipt / Notes'
ws_pend['G4'] = 'Status'

# Pre-fill Status column with "Pending" for rows 5-204
for row in range(5, 205):
    ws_pend[f'G{row}'] = 'Pending'

# =========================================================
# 8. KEY & INSTRUCTIONS - update stale text
# =========================================================
ws_ki = wb['Key & Instructions']
for row in ws_ki.iter_rows():
    for cell in row:
        if cell.value and 'Roster sheet (col B)' in str(cell.value):
            cell.value = "Set unique codes in the 'Access Codes' tab. Share each player's code only with them."
        if cell.value and 'Access Code' in str(cell.value) and 'col B' in str(cell.value):
            cell.value = str(cell.value).replace('col B', 'Access Codes tab')

wb.save(OUT)
print('Saved:', OUT)
