import openpyxl
from openpyxl.utils import get_column_letter
import sys

path = sys.argv[1]
sheet_name = sys.argv[2] if len(sys.argv) > 2 else None
max_row = int(sys.argv[3]) if len(sys.argv) > 3 else 12

wb = openpyxl.load_workbook(path)
sheets = [sheet_name] if sheet_name else wb.sheetnames

for name in sheets:
    ws = wb[name]
    print(f'\n========== {name} ==========')
    for row in ws.iter_rows(min_row=1, max_row=min(max_row, ws.max_row)):
        for cell in row:
            if cell.value is None and not cell.has_style:
                continue
            try:
                val = str(cell.value)[:40] if cell.value else '<empty>'
                val = val.encode('ascii', 'replace').decode()
            except Exception:
                val = '<encode-err>'
            f = cell.font
            fi = cell.fill
            al = cell.alignment
            b = cell.border
            try:
                font_color = f.color.rgb if f.color and f.color.type == 'rgb' else f'theme={f.color.theme}'
            except Exception:
                font_color = '?'
            try:
                fill_color = fi.fgColor.rgb if fi.fgColor and fi.fgColor.type == 'rgb' else f'theme={fi.fgColor.theme}'
            except Exception:
                fill_color = '?'
            print(
                f'  {cell.coordinate:6} | val={val:40} | '
                f'bold={str(f.bold):5} color={font_color} size={f.size} name={f.name} | '
                f'fill={fi.fill_type} fg={fill_color} | '
                f'halign={al.horizontal} | '
                f'border_top={b.top.border_style} bot={b.bottom.border_style}'
            )
