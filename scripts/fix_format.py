import shutil
from copy import copy
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SRC = r'C:\Users\jakep\AppData\Local\Temp\finance_data_modified.xlsx'
OUT = r'C:\Users\jakep\AppData\Local\Temp\finance_data_v3.xlsx'
shutil.copy(SRC, OUT)

wb = load_workbook(OUT)

# ── Shared style builders ─────────────────────────────────────────
NAVY      = '001A3A5C'
WHITE     = '00FFFFFF'
BLACK     = '00000000'
LT_BLUE   = '00D6E4F0'   # summary header fill
BLUE      = '000000FF'   # input / cross-sheet formula text
GREEN     = '00008000'   # cross-sheet formula text
AMBER     = '00B8860B'   # attendance input-row label
LT_YELLOW = '00FFF8DC'   # attendance input-row fill
LT_GRAY   = '00F5F5F5'   # empty data-row fill
GRAY_TEXT = '00AAAAAA'   # access-code / muted text

thin = Side(style='thin')
THIN_BORDER  = Border(top=thin, bottom=thin, left=thin, right=thin)
TB_BORDER    = Border(top=thin, bottom=thin)          # top+bottom only (for header spans)

def font(bold=False, color=BLACK, size=10, name='Arial'):
    return Font(bold=bold, color=color, size=size, name=name)

def fill(rgb):
    return PatternFill('solid', fgColor=rgb)

def align(h='left', v='center', wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def style(cell, *, bold=False, fc=BLACK, size=10, bg=None, h='left', border=True):
    cell.font      = font(bold=bold, color=fc, size=size)
    if bg:
        cell.fill  = fill(bg)
    cell.alignment = align(h)
    if border:
        cell.border = THIN_BORDER

def clear_style(cell):
    cell.font      = Font(name='Arial', size=10)
    cell.fill      = PatternFill()
    cell.alignment = Alignment()
    cell.border    = Border()

# ── Style sets (apply to a cell) ─────────────────────────────────
def title(cell):
    style(cell, bold=True, fc=WHITE, size=14, bg=NAVY, h='center')

def col_header(cell):
    style(cell, bold=True, fc=WHITE, bg=NAVY, h='center')

def summary_header(cell):
    style(cell, bold=True, fc=NAVY, bg=LT_BLUE, h='center')

def summary_data(cell):
    style(cell, fc=BLACK, h='right')

def input_label(cell):       # amber label in attendance input rows
    style(cell, bold=True, fc=AMBER, bg=LT_YELLOW, h='left')

def input_data(cell):        # blue cells in attendance input rows
    style(cell, fc=BLUE, bg=LT_YELLOW, h='left')

def data_text(cell, color=BLACK, h='left'):
    style(cell, fc=color, h=h)

def data_formula(cell):      # within-sheet formula (black)
    style(cell, fc=BLACK, h='right')

def xsheet_formula(cell):    # cross-sheet formula (green)
    style(cell, fc=GREEN, h='left')

def att_total_label(cell):
    style(cell, bold=True, fc=NAVY, bg=LT_BLUE, h='left')

def att_total_data(cell):
    style(cell, fc=BLACK, bg=LT_BLUE, h='right')

def empty_row_cell(cell, xsheet_col=False):
    style(cell, fc=(GREEN if xsheet_col else BLACK), bg=LT_GRAY)

# ── Helpers ──────────────────────────────────────────────────────
def apply_row(ws, row, col_start, col_end, fn):
    for c in range(col_start, col_end + 1):
        fn(ws.cell(row=row, column=c))

def clear_row_style(ws, row, col_start, col_end):
    for c in range(col_start, col_end + 1):
        clear_style(ws.cell(row=row, column=c))

# ─────────────────────────────────────────────────────────────────
# 1. ROSTER
# ─────────────────────────────────────────────────────────────────
ws = wb['Roster']
MAX_COL = 6   # A-F after cleanup

# Row 1: title (was already styled but lost navy fill on re-written A1)
title(ws['A1'])
for c in range(2, MAX_COL + 1):
    cell = ws.cell(row=1, column=c)
    cell.fill   = fill(NAVY)
    cell.border = THIN_BORDER

# Row 4: column headers (was row 6 before delete — keeps formatting from shift)
# Re-apply to be safe; new "Dues Amount ($)" label is the only changed header
for c in range(1, MAX_COL + 1):
    col_header(ws.cell(row=4, column=c))

# Data rows 5-204
for r in range(5, 205):
    a = ws.cell(row=r, column=1)
    if a.value is not None and str(a.value).strip():
        # Filled data rows
        data_text(a)                                         # A: player name, black
        data_text(ws.cell(row=r, column=2), color=BLUE)     # B: dues amount, blue (input)
        for c in range(3, MAX_COL + 1):
            data_text(ws.cell(row=r, column=c))
    else:
        # Empty placeholder rows
        for c in range(1, MAX_COL + 1):
            empty_row_cell(ws.cell(row=r, column=c))

# ─────────────────────────────────────────────────────────────────
# 2. EVENTS
# ─────────────────────────────────────────────────────────────────
ws = wb['Events']
MAX_COL = 10  # A-J after cleanup

# Row 1: title
title(ws['A1'])
for c in range(2, MAX_COL + 1):
    cell = ws.cell(row=1, column=c)
    cell.fill   = fill(NAVY)
    cell.border = THIN_BORDER

# Row 4: column headers (was row 6; End Date col C is new, rest preserved)
for c in range(1, MAX_COL + 1):
    col_header(ws.cell(row=4, column=c))

# Data rows 5-504: apply formatting to all columns
# Cols A, D, E, H, I, J = text (black), cols B, C = dates (black), cols F, G = numbers (blue input)
INPUT_COLS = {6, 7}   # F=Budgeted, G=Actual
DATE_COLS  = {2, 3}   # B=Start Date, C=End Date

for r in range(5, 505):
    a = ws.cell(row=r, column=1)
    is_filled = a.value is not None and str(a.value).strip()
    for c in range(1, MAX_COL + 1):
        cell = ws.cell(row=r, column=c)
        if is_filled:
            if c in INPUT_COLS:
                data_text(cell, color=BLUE)
            else:
                data_text(cell)
        else:
            empty_row_cell(cell)

# End Date column: needs a slightly wider default width
ws.column_dimensions['C'].width = 14

# ─────────────────────────────────────────────────────────────────
# 3. ATTENDANCE  (full rebuild — no existing styles)
# ─────────────────────────────────────────────────────────────────
ws = wb['Attendance']
N_EVENTS = 50      # cols B-AY
LAST_COL  = 51     # AY = col 51
N_PLAYERS = 200    # rows 7-206

# Row 1: title
title(ws['A1'])
for c in range(2, LAST_COL + 1):
    cell = ws.cell(row=1, column=c)
    cell.fill   = fill(NAVY)
    cell.border = THIN_BORDER

# Row 3: column headers (Player Name + event names from Events)
col_header(ws.cell(row=3, column=1))
for c in range(2, LAST_COL + 1):
    col_header(ws.cell(row=3, column=c))

# Row 4: cost per person (input row)
input_label(ws.cell(row=4, column=1))
for c in range(2, LAST_COL + 1):
    input_data(ws.cell(row=4, column=c))

# Row 5: Total Attendees
att_total_label(ws.cell(row=5, column=1))
for c in range(2, LAST_COL + 1):
    att_total_data(ws.cell(row=5, column=c))

# Row 6: Total Cost
att_total_label(ws.cell(row=6, column=1))
for c in range(2, LAST_COL + 1):
    att_total_data(ws.cell(row=6, column=c))

# Rows 7-206: player data
for r in range(7, 207):
    # Col A: player name (formula from Roster → green cross-sheet)
    xsheet_formula(ws.cell(row=r, column=1))
    ws.cell(row=r, column=1).alignment = align('left')
    # Cols B-AY: attendance marks (blue input, center)
    for c in range(2, LAST_COL + 1):
        style(ws.cell(row=r, column=c), fc=BLUE, h='center')

# Column widths for Attendance
ws.column_dimensions['A'].width = 22
for i in range(N_EVENTS):
    col = get_column_letter(i + 2)
    ws.column_dimensions[col].width = 22   # event name cols need room

# ─────────────────────────────────────────────────────────────────
# 4. PLAYER BALANCES
# ─────────────────────────────────────────────────────────────────
ws = wb['Player Balances']
MAX_COL = 6   # A-F after restructure

# Row 1: title
title(ws['A1'])
for c in range(2, MAX_COL + 1):
    cell = ws.cell(row=1, column=c)
    cell.fill   = fill(NAVY)
    cell.border = THIN_BORDER

# Row 3: summary headers
for c in range(1, 6):
    summary_header(ws.cell(row=3, column=c))
for c in range(6, MAX_COL + 1):
    clear_style(ws.cell(row=3, column=c))

# Row 4: summary data
for c in range(1, 6):
    summary_data(ws.cell(row=4, column=c))
for c in range(6, MAX_COL + 1):
    clear_style(ws.cell(row=4, column=c))

# Row 6: column headers
for c in range(1, MAX_COL + 1):
    col_header(ws.cell(row=6, column=c))

# Data rows 7-206
XSHEET_COLS = {1, 2, 3, 4}   # A (from Roster), B (Total Charged), C (Total Paid), D (Paid Upfront)
FORMULA_COLS = {5, 6}          # E (Net Balance), F (Status)

for r in range(7, 207):
    a = ws.cell(row=r, column=1)
    has_name = a.value is not None and str(a.value).strip()
    for c in range(1, MAX_COL + 1):
        cell = ws.cell(row=r, column=c)
        if has_name or a.value is not None:  # formula row (even if empty display)
            if c in XSHEET_COLS:
                xsheet_formula(cell)
                if c == 1:
                    cell.alignment = align('left')
            else:
                data_formula(cell)
        else:
            empty_row_cell(cell, xsheet_col=(c in XSHEET_COLS))

# Clear out any leftover cols G-H
for r in range(1, 207):
    for c in range(7, 9):
        clear_style(ws.cell(row=r, column=c))

# Column widths
ws.column_dimensions['A'].width = 22
ws.column_dimensions['B'].width = 20
ws.column_dimensions['C'].width = 18
ws.column_dimensions['D'].width = 24
ws.column_dimensions['E'].width = 18
ws.column_dimensions['F'].width = 18

# ─────────────────────────────────────────────────────────────────
# 5. PENDING APPROVAL  (new sheet — needs full styling)
# ─────────────────────────────────────────────────────────────────
ws = wb['Pending Approval']
MAX_COL = 7   # A-G

# Row 1: title
title(ws['A1'])
for c in range(2, MAX_COL + 1):
    cell = ws.cell(row=1, column=c)
    cell.fill   = fill(NAVY)
    cell.border = THIN_BORDER

# Row 2: instruction text
ws['A2'].font      = Font(name='Arial', size=10, italic=True, color='00555555')
ws['A2'].alignment = align('left')

# Row 4: column headers
for c in range(1, MAX_COL + 1):
    col_header(ws.cell(row=4, column=c))

# Rows 5-204: data rows
STATUS_COL = 7
for r in range(5, 205):
    for c in range(1, MAX_COL + 1):
        cell = ws.cell(row=r, column=c)
        if c == STATUS_COL:
            # Status pre-filled "Pending" — keep it visible
            data_text(cell, color='00888888', h='center')
        else:
            style(cell, fc=BLUE, h='left')  # input cells, blue

# Column widths
ws.column_dimensions['A'].width = 14   # Date
ws.column_dimensions['B'].width = 22   # Player Name
ws.column_dimensions['C'].width = 30   # Description
ws.column_dimensions['D'].width = 16   # Category
ws.column_dimensions['E'].width = 14   # Amount
ws.column_dimensions['F'].width = 30   # Receipt/Notes
ws.column_dimensions['G'].width = 14   # Status

# ─────────────────────────────────────────────────────────────────
# 6. DASHBOARD — clean up title cell (re-written A1 lost style)
# ─────────────────────────────────────────────────────────────────
ws = wb['Dashboard']
# Title was not changed, so style should be intact; just ensure it is
if ws['A1'].font.bold is not True:
    title(ws['A1'])
    for c in range(2, 3):
        ws.cell(row=1, column=c).fill   = fill(NAVY)
        ws.cell(row=1, column=c).border = THIN_BORDER

wb.save(OUT)
print('Saved:', OUT)
