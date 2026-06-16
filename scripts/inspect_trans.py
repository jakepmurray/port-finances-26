import openpyxl
wb = openpyxl.load_workbook(r'C:\Users\jakep\OneDrive\Documents\Port Authority\2026\port-finances-26\finance_data.xlsx')

print("=== Events cols (row 4 headers + row 5 data) ===")
ws = wb['Events']
for r in [4, 5, 6]:
    row = [str(ws.cell(r, c).value)[:25] if ws.cell(r, c).value else '-' for c in range(1, 11)]
    print(f"  row{r}: {row}")

print("\n=== Attendance row 3 (first 6 cols) and row 4 ===")
ws = wb['Attendance']
for r in [3, 4, 5]:
    row = [str(ws.cell(r, c).value)[:30] if ws.cell(r, c).value else '-' for c in range(1, 7)]
    print(f"  row{r}: {row}")

print("\n=== Transactions (rows 6-14) ===")
ws = wb['Transactions']
for r in range(6, 15):
    row = [str(ws.cell(r, c).value)[:20] if ws.cell(r, c).value else '-' for c in range(1, 10)]
    print(f"  row{r}: {row}")

print("\n=== Player Balances (rows 6-10) ===")
ws = wb['Player Balances']
for r in range(6, 11):
    row = [str(ws.cell(r, c).value)[:35] if ws.cell(r, c).value else '-' for c in range(1, 7)]
    print(f"  row{r}: {row}")
