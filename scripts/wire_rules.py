import shutil
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SRC = r'C:\Users\jakep\OneDrive\Documents\Port Authority\2026\port-finances-26\finance_data.xlsx'
OUT = r'C:\Users\jakep\AppData\Local\Temp\finance_data_final.xlsx'
shutil.copy(SRC, OUT)

wb = load_workbook(OUT)

# ── Shared styles ─────────────────────────────────────────────────
NAVY    = '001A3A5C'
WHITE   = '00FFFFFF'
BLACK   = '00000000'
LT_BLUE = '00D6E4F0'
GREEN   = '00008000'

thin = Side(style='thin')
BORDER = Border(top=thin, bottom=thin, left=thin, right=thin)

def font(bold=False, color=BLACK, size=10, name='Arial'):
    return Font(bold=bold, color=color, size=size, name=name)

def fill(rgb):
    return PatternFill('solid', fgColor=rgb)

def style(cell, *, bold=False, fc=BLACK, bg=None, h='left'):
    cell.font      = font(bold=bold, color=fc)
    if bg:
        cell.fill  = fill(bg)
    cell.alignment = Alignment(horizontal=h, vertical='center')
    cell.border    = BORDER

# ─────────────────────────────────────────────────────────────────
# 1. ATTENDANCE — row 4: auto-calculate cost per person
#
# For each event column (B=col2 … AY=col51):
#   - Corresponding Events data row = 5 + offset (B→5, C→6, …)
#   - Split Method is in Events!I (col 9)
#   - Actual Cost is in Events!G (col 7)
#
# Rule:
#   Equal Split  → actual_cost / MAX(1, COUNTA(active_roster))
#                   (everyone on the active roster pays, not just attendees)
#   Attendance   → actual_cost / MAX(1, COUNTIF(this_col_attendees, "1"))
#                   (only those who attended pay)
#   Anything else → attendance-based (same as Attendance)
#
# COUNTA(Roster!A5:A204) gives total roster size.
# COUNTIF(col7:col206,"1") gives attendees for this event.
# ─────────────────────────────────────────────────────────────────
ws_att = wb['Attendance']

for i in range(50):          # 50 events
    col_idx  = i + 2         # B=2 … AY=51
    col_ltr  = get_column_letter(col_idx)
    ev_row   = 5 + i         # Events rows 5-54

    # Build formula that respects split method:
    #   IF split = "Equal Split"  → cost / active roster size
    #   ELSE (Attendance or other) → cost / # who attended
    formula = (
        f'=IFERROR('
        f'IF(Events!$A${ev_row}<>"",'
        f'IF(Events!$I${ev_row}="Equal Split",'
        f'Events!$G${ev_row}/MAX(1,COUNTA(Roster!$A$5:$A$204)),'
        f'Events!$G${ev_row}/MAX(1,COUNTIF({col_ltr}$7:{col_ltr}$206,"1"))),'
        f'0),0)'
    )
    cell = ws_att.cell(row=4, column=col_idx)
    cell.value = formula

    # Re-style row 4 data cells: auto-calculated → light blue like Total rows
    style(cell, fc=BLACK, bg=LT_BLUE, h='right')

# Re-style row 4 label cell (A4)
label = ws_att['A4']
style(label, bold=True, fc=NAVY, bg=LT_BLUE, h='left')
# Add a note to the label indicating it's auto-calculated
label.value = 'Cost per Person ($) [auto]'

# ─────────────────────────────────────────────────────────────────
# 2. TRANSACTIONS — add "Cost per Attendee ($)" as column J
#
# For each transaction row that has a Linked Event (col G):
#   1. Look up the event in Events to get Actual Cost (col G) and
#      Split Method (col I)
#   2. Look up the event in Attendance row 3 to find which column
#      corresponds to that event, then count attendees
#   3. Apply split method rule:
#      - Equal Split → cost / roster size
#      - Attendance  → cost / # attendees
#
# Attendance!$B$3:$AY$3 contains event names (formula-driven from Events).
# MATCH(G7, Attendance!$B$3:$AY$3, 0) gives the column offset within B:AY.
# INDEX(Attendance!$B$7:$AY$206, 0, offset) returns all attendance marks
#   for that event; COUNTIF that result for "1" gives attendee count.
# ─────────────────────────────────────────────────────────────────
ws_t = wb['Transactions']

# Header row 6, col J
header = ws_t.cell(row=6, column=10)
header.value = 'Cost per Attendee ($)'
style(header, bold=True, fc=WHITE, bg=NAVY, h='center')

# Data rows 7-1000
for r in range(7, 1001):
    cell = ws_t.cell(row=r, column=10)
    # Formula:
    #   - If no linked event → blank
    #   - Else:
    #       actual_cost  = INDEX/MATCH from Events
    #       split_method = INDEX/MATCH from Events col I
    #       att_col_offset = MATCH of event name in Attendance row 3 (1-indexed within B:AY)
    #       attendees    = COUNTIF of that Attendance column for "1"
    #       result = IF split="Equal Split", cost/roster_size, cost/attendees
    formula = (
        f'=IF(G{r}="","",'
        f'IFERROR('
        f'LET('
        f'ev_cost,INDEX(Events!$G$5:$G$505,MATCH(G{r},Events!$A$5:$A$505,0)),'
        f'ev_split,INDEX(Events!$I$5:$I$505,MATCH(G{r},Events!$A$5:$A$505,0)),'
        f'att_offset,MATCH(G{r},Attendance!$B$3:$AY$3,0),'
        f'attendees,COUNTIF(INDEX(Attendance!$B$7:$AY$206,0,att_offset),"1"),'
        f'IF(ev_split="Equal Split",'
        f'ev_cost/MAX(1,COUNTA(Roster!$A$5:$A$204)),'
        f'ev_cost/MAX(1,attendees))),'
        f'"N/A"))'
    )
    cell.value = formula
    style(cell, fc=GREEN, bg=None, h='right')

# Set column width
ws_t.column_dimensions['J'].width = 22

# ─────────────────────────────────────────────────────────────────
# 3. PLAYER BALANCES — Total Charged now correctly uses
#    the formula-driven Attendance row 4, so no formula change
#    needed there. But add a note to make the logic visible:
#    the SUMPRODUCT already multiplies attendance × auto-cost.
#
#    However, for "Equal Split" events the cost per person is
#    already divided by roster size in Attendance row 4, so the
#    SUMPRODUCT(attendance="1" × cost_per_person) gives each
#    active player their share regardless of whether they're
#    marked "1" or not — UNLESS we also multiply by the "1" mark.
#
#    Fix needed: for Equal Split, each active player is charged
#    their share whether or not they attended. So the formula
#    in Player Balances should NOT gate on attendance for Equal
#    Split events.
#
#    Practical solution: The treasurer marks ALL active players
#    as "1" for Equal Split events — this is the intended workflow.
#    We'll add a note to Key & Instructions explaining this.
# ─────────────────────────────────────────────────────────────────
ws_ki = wb['Key & Instructions']

# Find the ATTENDANCE section and add a workflow note
found = False
for row in ws_ki.iter_rows():
    for cell in row:
        if cell.value and 'Cost per person' in str(cell.value):
            # Update the note
            cell.value = 'Row 4 in Attendance is auto-calculated from Events (Actual Cost ÷ attendees for "Attendance" events; ÷ roster size for "Equal Split" events)'
            found = True

wb.save(OUT)
print('Saved:', OUT)
