import openpyxl

wb = openpyxl.load_workbook(r'C:\Users\jakep\AppData\Local\Temp\finance_data_bugfix.xlsx', data_only=True)

# Print ALL rows 1-12 for key sheets (including empty ones), cols 1-10
for sheet_name in ['Roster', 'Player Balances']:
    ws = wb[sheet_name]
    print(f'\n=== {sheet_name} (ALL rows 1-12) ===')
    for r in range(1, 13):
        row = [ws.cell(r, c).value for c in range(1, 11)]
        print(f'  row{r}: {row}')

# Access Codes - print raw values safely
print('\n=== Access Codes (rows 1-10) ===')
ws = wb['Access Codes']
for r in range(1, 11):
    row = [str(ws.cell(r, c).value).encode('ascii','replace').decode() for c in range(1, 8)]
    if any(v != 'None' for v in row):
        print(f'  row{r}: {row}')
